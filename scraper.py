"""
Toronto Police Service – Crime-focused news-release scraper
‾‾‾‾‾‾‾‾‾‾‾‾‾‾‾‾‾‾‾‾‾‾‾‾‾‾‾‾‾‾‾‾‾‾‾‾‾‾‾‾‾‾
 • Headless-Chrome via Selenium (bypasses 403 / JS)
 • Scans N recent releases (default 20)
 • Filters articles that mention BOTH:
      – any crime keyword
      – an age pattern (“22-year-old”, “22 year old”)
 • Extracts Name • Age • Crime from narrative text
 • Saves results to crime_data_final.xlsx (Excel-ready)
"""

import os
import re
import time
import pandas as pd
import openpyxl
import chromedriver_autoinstaller
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from bs4 import BeautifulSoup
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

BASE = "https://www.tps.ca"

# ──────────────────────────────────────
# 1. Crime keyword list
# ──────────────────────────────────────
CRIME_KEYWORDS = [
    # violent
    "murder", "attempted murder", "homicide", "manslaughter", "assault",
    "aggravated assault", "assault with a weapon", "assault causing bodily harm",
    "uttering threats", "forcible confinement", "criminal negligence causing death",
    "attempted strangulation",
    # firearms / weapons
    "firearm", "weapon", "handgun", "rifle", "gun", "shotgun", "discharge firearm",
    "possess loaded firearm", "carry concealed weapon", "unauthorized possession",
    "use of a firearm in crime",
    # property
    "robbery", "break and enter", "burglary", "theft", "auto theft",
    "possession of stolen property", "arson", "mischief under", "mischief over",
    "vandalism", "trespassing", "tampering",
    # drugs
    "trafficking", "possession for the purpose", "controlled substance", "cocaine",
    "heroin", "fentanyl", "methamphetamine", "marijuana", "illicit drugs", "drug lab",
    # driving / dui
    "dui", "impaired driving", "driving under the influence", "over 80",
    "refuse breath sample", "blood alcohol concentration", "drug-impaired driving",
    "operating while impaired", "dangerous driving", "fail to remain", "evading police",
    "reckless driving", "stunt driving", "high-speed pursuit", "police chase",
    "criminal negligence in operation of a vehicle",
    # sexual
    "sexual assault", "sexual interference", "invitation to sexual touching",
    "child luring", "indecent exposure", "pornography", "voyeurism", "internet luring",
    # organized / other
    "gang-related", "hate crime", "human trafficking", "extortion", "intimidation",
    "criminal organization", "fraud", "financial crime",
    # procedural
    "breach of probation", "fail to comply", "obstruct police", "resist arrest",
    "escape lawful custody", "perjury", "public mischief", "impersonation of police",
    # narrative helpers
    "charged with", "arrested for", "suspected of", "wanted for", "under investigation for"
]
CRIME_SET = {kw.lower() for kw in CRIME_KEYWORDS}

# ──────────────────────────────────────
# 2. Start Chrome (headless & safe for cloud)
# ──────────────────────────────────────
def start_driver():
    chromedriver_autoinstaller.install()

    opts = Options()
    opts.headless = True
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_argument("--window-size=1920,1080")

    return webdriver.Chrome(options=opts)

# ──────────────────────────────────────
# 3. Get recent news release links
# ──────────────────────────────────────
def recent_links(driver, max_links=100):
    links = []
    page = 1
    while len(links) < max_links:
        url = f"{BASE}/media-centre/news-releases/?page={page}"
        driver.get(url)
        time.sleep(3)
        soup = BeautifulSoup(driver.page_source, "html.parser")
        new = []
        for a in soup.select('a[href^="/media-centre/news-releases/"]'):
            slug = a["href"].rstrip("/").split("/")[-1]
            if not slug.isdigit():
                continue
            url_full = BASE + a["href"]
            if url_full in [l["url"] for l in links]:
                continue
            new.append({
                "url": url_full,
                "title": a.get_text(strip=True) or "No title",
                "date": a.find_next("time").get_text(strip=True) if a.find_next("time") else ""
            })
        if not new:
            break
        links.extend(new)
        print(f"Page {page}: found {len(new)} links (total {len(links)})")
        page += 1
    return links[:max_links]

# ──────────────────────────────────────
# 4. Extract full article text
# ──────────────────────────────────────
def article_text(driver, url):
    driver.get(url)
    time.sleep(1.5)
    soup = BeautifulSoup(driver.page_source, "html.parser")
    zone = soup.select_one("div.grid-container") or soup.find("article") or soup
    full_text = zone.get_text(" ", strip=True)
    pub_date_match = re.search(r"Published:\s*(.*?\d{4})", full_text)
    pub_date = pub_date_match.group(1).strip() if pub_date_match else ""
    return full_text, pub_date

# ──────────────────────────────────────
# 5. Filter for crime-related content
# ──────────────────────────────────────
def is_crime_related(text):
    return any(k in text.lower() for k in CRIME_SET)

# ──────────────────────────────────────
# 6. Suspect extraction
# ──────────────────────────────────────
SUSPECT_BLOCK = re.compile(
    r"([A-Z][\w'’\-]+(?: [A-Z][\w'’\-]+){0,2}),\s*(\d{1,3}).{0,80}?"
    r"(murder|assault|homicide|robbery|theft|firearm|weapon|traffick|dui|impaired|sexual|drug|dangerous driving|fail to remain)",
    re.IGNORECASE
)

def extract_suspects(text):
    hits = [
        {"Name": n.strip(), "Age": a.strip(), "Crime": c.strip()}
        for n, a, c in SUSPECT_BLOCK.findall(text)
    ]
    if hits:
        return hits

    # Fallback pattern
    crime = re.search(r"(?:charged with|arrested for|suspected of|wanted for)\s+(.{5,80}?)\.", text, re.IGNORECASE)
    if not crime:
        return []
    name = re.search(r"\b([A-Z][a-z]+(?: [A-Z][a-z]+){1,2})\b", text)
    age = re.search(r"\b(\d{1,3})\s*(?:year|yrs)[-\s]?old\b", text, re.IGNORECASE)
    return [{
        "Name": name.group(1) if name else None,
        "Age": age.group(1) if age else None,
        "Crime": crime.group(1).strip()
    }]

# ──────────────────────────────────────
# 7. Main function
# ──────────────────────────────────────
def main(max_links=20):
    driver = start_driver()
    rows = []
    seen_keys = set()
    try:
        for link in recent_links(driver, max_links):
            print(f"\n🔎 {link['title']}")
            txt, pub_date = article_text(driver, link["url"])
            if not is_crime_related(txt):
                print("   – not crime-related, skipped")
                continue
            suspects = extract_suspects(txt)
            if not suspects:
                print("   – crime article but no suspect pattern found")
                continue
            print(f"   – {len(suspects)} suspect(s) extracted")
            for s in suspects:
                key = (s["Name"], s["Age"], link["url"])
                if key in seen_keys:
                    continue
                seen_keys.add(key)
                rows.append({
                    "Name": s["Name"],
                    "Age": s["Age"],
                    "Crime": s["Crime"],
                    "Article": link["url"],
                    "Date": pub_date
                })
    finally:
        driver.quit()

    df = pd.DataFrame(rows)
    for col in ("date", "title", "Title"):
        if col in df.columns:
            df.drop(columns=[col], inplace=True)

    excel_file = "crime_data_final.xlsx"
    if os.path.exists(excel_file):
        os.remove(excel_file)

    df.to_excel(excel_file, index=False)
    wb = load_workbook(excel_file)
    ws = wb.active

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_len + 2

    ws.freeze_panes = "A2"
    wb.save(excel_file)
    print(f"\n✅ Saved {len(rows)} rows ➜ {excel_file}")

# ──────────────────────────────────────
if __name__ == "__main__":
    main(200)
