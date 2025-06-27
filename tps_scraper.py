"""
Toronto Police Service – Crime-focused news-release scraper

 • Headless Chrome via Selenium
 • Scans recent news releases (default: 20)
 • Filters for articles with:
     – crime keywords
     – age patterns (e.g. “22-year-old”)
 • Extracts Name • Age • Crime
 • Outputs Excel: crime_data_final.xlsx
"""

import os
import re
import time
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

# ──────────────────────────────────────
# Configuration
# ──────────────────────────────────────
BASE_URL = "https://www.tps.ca"

CRIME_KEYWORDS = [
    "murder", "homicide", "manslaughter", "assault", "weapon", "firearm",
    "gun", "robbery", "break and enter", "trafficking", "sexual assault",
    "dui", "impaired driving", "gang", "fraud", "human trafficking",
    "hate crime", "child luring", "pornography", "dangerous driving",
    "stunt driving", "public mischief", "wanted for", "arrested for",
    "charged with", "under investigation"
]
CRIME_SET = {kw.lower() for kw in CRIME_KEYWORDS}

SUSPECT_PATTERN = re.compile(
    r"([A-Z][\w'’\-]+(?: [A-Z][\w'’\-]+){0,2}),\s*(\d{1,3}).{0,80}?"
    r"(murder|assault|homicide|robbery|theft|firearm|weapon|traffick|dui|impaired|sexual|drug|driving|fail to remain)",
    re.IGNORECASE
)

# ──────────────────────────────────────
# Start Chrome driver
# ──────────────────────────────────────
def start_driver():
    options = Options()
    options.binary_location = "/usr/bin/google-chrome"
    options.add_argument("--headless")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--window-size=1920,1080")

    # Use system-installed chromedriver
    service = Service("/usr/bin/chromedriver")

    return webdriver.Chrome(service=service, options=options)
# ──────────────────────────────────────
# Get recent article links
# ──────────────────────────────────────
def get_recent_links(driver, max_links=20):
    links = []
    page = 1
    while len(links) < max_links:
        driver.get(f"{BASE_URL}/media-centre/news-releases/?page={page}")
        time.sleep(2)
        soup = BeautifulSoup(driver.page_source, "html.parser")
        new_links = []

        for a in soup.select('a[href^="/media-centre/news-releases/"]'):
            slug = a["href"].rstrip("/").split("/")[-1]
            if not slug.isdigit():
                continue
            full_url = BASE_URL + a["href"]
            if full_url in [l["url"] for l in links]:
                continue
            new_links.append({
                "url": full_url,
                "title": a.get_text(strip=True) or "No title"
            })

        if not new_links:
            break
        links.extend(new_links)
        page += 1
    return links[:max_links]

# ──────────────────────────────────────
# Extract full article text and date
# ──────────────────────────────────────
def get_article_text(driver, url):
    driver.get(url)
    time.sleep(1.5)
    soup = BeautifulSoup(driver.page_source, "html.parser")
    zone = soup.select_one("div.grid-container") or soup.find("article") or soup
    text = zone.get_text(" ", strip=True)
    date_match = re.search(r"Published:\s*(.*?\d{4})", text)
    pub_date = date_match.group(1).strip() if date_match else ""
    return text, pub_date

# ──────────────────────────────────────
# Crime article check
# ──────────────────────────────────────
def is_crime_related(text):
    return any(word in text.lower() for word in CRIME_SET)

# ──────────────────────────────────────
# Extract suspect info
# ──────────────────────────────────────
def extract_suspects(text):
    matches = SUSPECT_PATTERN.findall(text)
    if matches:
        return [{"Name": n, "Age": a, "Crime": c} for n, a, c in matches]

    fallback_crime = re.search(r"(?:charged with|arrested for|suspected of|wanted for)\s+(.+?)\.", text, re.IGNORECASE)
    if fallback_crime:
        name = re.search(r"\b([A-Z][a-z]+(?: [A-Z][a-z]+){1,2})\b", text)
        age = re.search(r"\b(\d{1,3})\s*(?:year|yrs)[-\s]?old\b", text, re.IGNORECASE)
        return [dict(
            Name=name.group(1) if name else None,
            Age=age.group(1) if age else None,
            Crime=fallback_crime.group(1).strip()
        )]
    return []

# ──────────────────────────────────────
# Main function
# ──────────────────────────────────────
def main(max_links=20):
    driver = start_driver()
    rows = []
    seen = set()
    try:
        for link in get_recent_links(driver, max_links):
            print(f"\n🔍 {link['title']}")
            text, date = get_article_text(driver, link["url"])
            if not is_crime_related(text):
                print("   – Skipped (not crime-related)")
                continue
            suspects = extract_suspects(text)
            if not suspects:
                print("   – Crime article but no suspects found")
                continue
            print(f"   – Found {len(suspects)} suspect(s)")
            for s in suspects:
                key = (s["Name"], s["Age"], link["url"])
                if key in seen:
                    continue
                seen.add(key)
                rows.append({
                    "Name": s["Name"],
                    "Age": s["Age"],
                    "Crime": s["Crime"],
                    "Article": link["url"],
                    "Date": date
                })
    finally:
        driver.quit()

    if rows:
        df = pd.DataFrame(rows)
        output_file = "crime_data_final.xlsx"
        df.to_excel(output_file, index=False)

        wb = load_workbook(output_file)
        ws = wb.active
        for col in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2
        ws.freeze_panes = "A2"
        wb.save(output_file)
        print(f"\n✅ Saved {len(rows)} records to ➜ {output_file}")
    else:
        print("⚠️ No valid crime data extracted.")

# ──────────────────────────────────────
if __name__ == "__main__":
    main(20)
